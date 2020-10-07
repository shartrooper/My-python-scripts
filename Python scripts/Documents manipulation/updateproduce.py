#! python3
# updateproduce.py - Corrects costs in produce sales spreadsheet.

import openpyxl

wb = openpyxl.load_workbook("produceSales.xlsx")
sheet = wb["Sheet"]

# The produce types and their updated prices
PRICE_UPDATES = {"Garlic": 3.07, "Celery": 1.19, "Lemon": 1.27}

# Loop through the rows and update the prices.
for rowNum in range(2, sheet.max_row):  # skip the first row
    produceName = sheet.cell(row=rowNum, column=1).value
    if produceName in PRICE_UPDATES:
        '''
        Notice how clean using PRICE_UPDATES makes the code.
        Only one if statement, rather than code like if produceName == 'Garlic': ,
        is necessary for every type of produce to update.
        '''
        sheet.cell(row=rowNum, column=2).value = PRICE_UPDATES[produceName]
wb.save("updatedProduceSales.xlsx")
