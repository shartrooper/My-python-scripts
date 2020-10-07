# cellInverter.py invert the row and column of the cells in the spreadsheet
# Get filename from command line and create new file with the inverted cells

from openpyxl import Workbook, load_workbook
import sys

# take N number argument from command line.
if len(sys.argv) > 1:
    # Get arguments from command line.
    filename = sys.argv[1:][0]
else:
    print("Please, introduce filename in command lines")
    sys.exit(1)
# Load wordBook filename
try:
    wb = load_workbook(filename, read_only=True)
    ws = wb.active
    wb1 = Workbook()
    ws1 = wb1.active
except Exception as err:
    print("An exception happened : \n" + str(err))
    sys.exit(1)
# copy/paste cells from loaded workbook to new file.
for i, row in enumerate(ws.values, start=1):
    for j, copyValue in enumerate(row, start=1):
        ws1.cell(row=j, column=i, value=copyValue)
#create and save on a new file.


wb1.save(f"inverted-{filename}")
print("Inverted Workbook created and saved in local dir!")
