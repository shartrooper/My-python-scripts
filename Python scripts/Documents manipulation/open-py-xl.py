import openpyxl
wb = openpyxl.load_workbook('example.xlsx')
type(wb)
<class 'openpyxl.workbook.workbook.Workbook'>
wb.sheetnames # The workbook's sheets' names.
['Sheet1', 'Sheet2', 'Sheet3']
sheet = wb['Sheet3'] # Get a sheet from the workbook.
#sheet
<Worksheet "Sheet3">
>type(sheet)
<class 'openpyxl.worksheet.worksheet.Worksheet'>
sheet.title # Get the sheet's title as a string.
'Sheet3'
anotherSheet = wb.active # Get the active sheet.
anotherSheet
<Worksheet "Sheet1">

sheet = wb['Sheet1'] # Get a sheet from the workbook.
sheet['A1'] # Get a cell from the sheet.
<Cell 'Sheet1'.A1>
sheet['A1'].value # Get the value from the cell.
datetime.datetime(2015, 4, 5, 13, 34, 2)
c = sheet['B1'] # Get another cell from the sheet.
c.value
'Apples'
# Get the row, column, and value from the cell.
'Row %s, Column %s is %s' % (c.row, c.column, c.value)
'Row 1, Column B is Apples'
'Cell %s is %s' % (c.coordinate, c.value)
'Cell B1 is Apples'
sheet['C1'].value
73

'''
Specifying a column by letter can be tricky to program, especially because after column Z, the columns start by using two letters: AA, AB, AC, and so on.
As an alternative, you can also get a cell using the sheet’s cell() method and passing integers for its row and column keyword arguments.
The first row or column integer is 1, not 0. Continue the interactive shell example by entering the following:
'''
sheet.cell(row=1, column=2)
<Cell 'Sheet1'.B1>
sheet.cell(row=1, column=2).value
'Apples'
>for i in range(1, 8, 2): # Go through every other row:
#By passing 2 for the range() function’s “step” parameter, you can get cells from every second row (in this case, all the odd-numbered rows).
    print(i, sheet.cell(row=i, column=2).value)
#1 Apples
#3 Pears
#5 Apples
#7 Strawberries

##Converting Between Column Letters and Numbers

from openpyxl.utils import get_column_letter, column_index_from_string
get_column_letter(1) # Translate column 1 to a letter.
'A'
get_column_letter(2)
'B'
get_column_letter(27)
'AA'
get_column_letter(900)
'AHP'
wb = openpyxl.load_workbook('example.xlsx')
sheet = wb['Sheet1']
get_column_letter(sheet.max_column)
'C'
column_index_from_string('A') # Get A's number.
1
column_index_from_string('AA')
27

##Getting Rows and Columns from the Sheets
'''
You can slice Worksheet objects to get all the Cell objects in a row, column, or rectangular area of the spreadsheet.
Then you can loop over all the cells in the slice.
'''
sheet = wb['Sheet1']
tuple(sheet['A1':'C3']) # Get all cells from A1 to C3.
((<Cell 'Sheet1'.A1>, <Cell 'Sheet1'.B1>, <Cell 'Sheet1'.C1>), (<Cell'Sheet1'.A2>, <Cell 'Sheet1'.B2>, <Cell 'Sheet1'.C2>), (<Cell 'Sheet1'.A3>,<Cell 'Sheet1'.B3>, <Cell 'Sheet1'.C3>))
for rowOfCellObjects in sheet['A1':'C3']:
    for cellObj in rowOfCellObjects:
        print(cellObj.coordinate, cellObj.value)
    print('--- END OF ROW ---')
'''
   A1 2015-04-05 13:34:02
   B1 Apples
   C1 73
   --- END OF ROW ---
   A2 2015-04-05 03:41:23
   B2 Cherries
   C2 85
   --- END OF ROW ---
   A3 2015-04-06 12:46:51
   B3 Pears
   C3 14
   --- END OF ROW ---
'''

'''
To access the values of cells in a particular row or column, you can also use a Worksheet object’s rows and columns attribute.
These attributes must be converted to lists with the list() function before you can use the square brackets and an index with them.
'''

sheet = wb.active
list(sheet.columns)[1] # Get second column's cells.
(<Cell 'Sheet1'.B1>, <Cell 'Sheet1'.B2>, <Cell 'Sheet1'.B3>, <Cell 'Sheet1'.B4>, <Cell 'Sheet1'.B5>, <Cell 'Sheet1'.B6>, <Cell 'Sheet1'.B7>)
for cellObj in list(sheet.columns)[1]:
        print(cellObj.value)
#Apples
#Cherries
#Pears
#Oranges
#Apples
#Bananas
#Strawberries


'''
Using the rows attribute on a Worksheet object will give you a tuple of tuples.
Each of these inner tuples represents a row, and contains the Cell objects in that row.
The columns attribute also gives you a tuple of tuples, with each of the inner tuples containing the Cell objects in a particular column.
For example.xlsx, since there are 7 rows and 3 columns, rows gives us a tuple of 7 tuples (each containing 3 Cell objects),
and columns gives us a tuple of 3 tuples (each containing 7 Cell objects).
'''

'''
Workbooks, Sheets, Cells

As a quick review, here’s a rundown of all the functions, methods, and data types involved in reading a cell out of a spreadsheet file:

    Import the openpyxl module.
    Call the openpyxl.load_workbook() function.
    Get a Workbook object.
    Use the active or sheetnames attributes.
    Get a Worksheet object.
    Use indexing or the cell() sheet method with row and column keyword arguments.
    Get a Cell object.
    Read the Cell object’s value attribute.
'''

##Setting the Font Style of Cells

wb = openpyxl.Workbook()
sheet = wb['Sheet']
italic24Font = Font(size=24, italic=True) # Create a font.
sheet['A1'].font = italic24Font # Apply the font to A1.
sheet['A1'] = 'Hello, world!'
wb.save('styles.xlsx')

## FONT OBJECTS


'''
Keyword argument
	

Data type
	

Description

name
	

String
	

The font name, such as 'Calibri' or 'Times New Roman'

size
	

Integer
	

The point size

bold
	

Boolean
	

True, for bold font

italic
	

Boolean
	

True, for italic font
'''

#You can call Font() to create a Font object and store that Font object in a variable. You then assign that variable to a Cell object’s font attribute.

from openpyxl.styles import Font
wb = openpyxl.Workbook()
sheet = wb['Sheet']

fontObj1 = Font(name='Times New Roman', bold=True)
sheet['A1'].font = fontObj1
sheet['A1'] = 'Bold Times New Roman'

fontObj2 = Font(size=24, italic=True)
sheet['B3'].font = fontObj2
sheet['B3'] = '24 pt Italic'

wb.save('styles.xlsx')

##Formulas

sheet = wb.active
sheet['A1'] = 200
sheet['A2'] = 300
sheet['A3'] = '=SUM(A1:A2)' # Set the formula.
wb.save('writeFormula.xlsx')

##Setting Row Height and Column Width

sheet = wb.active
sheet['A1'] = 'Tall row'
sheet['B2'] = 'Wide column'
# Set the height and width:
sheet.row_dimensions[1].height = 70
sheet.column_dimensions['B'].width = 20
wb.save('dimensions.xlsx')

##Merging and Unmerging Cells

sheet = wb.active
sheet.merge_cells('A1:D3') # Merge all these cells.
sheet['A1'] = 'Twelve cells merged together.'
sheet.merge_cells('C5:D5') # Merge these two cells.
sheet['C5'] = 'Two merged cells.'
wb.save('merged.xlsx')

sheet = wb.active
sheet.unmerge_cells('A1:D3') # Split these cells up.
sheet.unmerge_cells('C5:D5')
wb.save('merged.xlsx')

##Freezing Panes

'''
For spreadsheets too large to be displayed all at once,
it’s helpful to “freeze” a few of the top rows or leftmost columns onscreen.
'''

'''
In OpenPyXL, each Worksheet object has a freeze_panes attribute that can be set to a Cell object or a string of a cell’s coordinates.
Note that all rows above and all columns to the left of this cell will be frozen, but the row and column of the cell itself will not be frozen.

To unfreeze all panes, set freeze_panes to None or 'A1'.
'''

'''
freeze_panes setting
	

Rows and columns frozen

sheet.freeze_panes = 'A2'
	

Row 1

sheet.freeze_panes = 'B1'
	

Column A

sheet.freeze_panes = 'C1'
	

Columns A and B

sheet.freeze_panes = 'C2'
	

Row 1 and columns A and B

sheet.freeze_panes = 'A1' or sheet.freeze_panes = None
	

No frozen panes
'''
wb = openpyxl.load_workbook('produceSales.xlsx')
sheet = wb.active
sheet.freeze_panes = 'A2' # Freeze the rows above A2.
wb.save('freezeExample.xlsx')
# If you set the freeze_panes attribute to 'A2', row 1 will always be viewable, no matter where the user scrolls in the spreadsheet.
